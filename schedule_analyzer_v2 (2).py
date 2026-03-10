"""
班级课表空余时间统计分析工具 - 增强专业版 v2
新增功能:
  · 异步并行解析 (asyncio + ThreadPoolExecutor)
  · 磁盘缓存 (pickle, 基于文件哈希)
  · 逐行读取 Excel (openpyxl 流模式)，大幅降低内存占用
  · scipy 稀疏矩阵存储 occ_sum
  · 图表懒加载 (切换到对应标签才生成)
  · 支持 ZIP 压缩包 / 直接多文件上传
  · 导出: Markdown 报告 / Excel (多 Sheet) / CSV / 图表 PNG
  · matplotlib 中文字体自动检测 (Noto Sans CJK SC)
  · 增强可视化: 气泡图、箱线分布、分组柱状图、日历热力图
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import os, re, zipfile, tempfile, asyncio, hashlib, pickle, io, shutil, time
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Tuple

# ── third-party ───────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import matplotlib
matplotlib.use("Agg")          # 无 GUI 后端
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from openpyxl import load_workbook
from scipy import sparse

# xlrd 是可选依赖，仅读取 .xls 时需要
try:
    import xlrd
    _XLRD_OK = True
except ImportError:
    _XLRD_OK = False

# ── 页面配置 ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="📅 课表空余分析",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# 中文字体配置（matplotlib）
# ─────────────────────────────────────────────────────────────────────────────
_CN_FONT = None

def _setup_cn_font():
    """自动检测并配置 matplotlib 中文字体"""
    global _CN_FONT
    if _CN_FONT is not None:
        return _CN_FONT
    candidates = [
        "Noto Sans CJK SC", "Noto Sans CJK TC", "Noto Sans CJK JP",
        "Noto Serif CJK SC", "Noto Serif CJK JP",
        "SimHei", "WenQuanYi Micro Hei", "Microsoft YaHei",
        "PingFang SC", "Heiti SC",
    ]
    available = {f.name for f in fm.fontManager.ttflist}
    for c in candidates:
        if c in available:
            _CN_FONT = c
            break
    if _CN_FONT is None:
        _CN_FONT = "DejaVu Sans"  # 降级方案（中文会显示方块，但不崩溃）
    matplotlib.rcParams["font.family"] = [_CN_FONT, "DejaVu Sans"]
    matplotlib.rcParams["axes.unicode_minus"] = False
    return _CN_FONT

_setup_cn_font()

# ─────────────────────────────────────────────────────────────────────────────
# 自定义 CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.main-header {
    font-size: 2.4rem; font-weight: 700;
    background: linear-gradient(135deg, #1E3A8A, #3B82F6);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    margin-bottom: .4rem;
}
.sub-header { font-size: 1rem; color: #6B7280; margin-bottom: 1.5rem; }

.kpi-card {
    background: linear-gradient(135deg, #EFF6FF, #DBEAFE);
    border-left: 4px solid #3B82F6;
    border-radius: .6rem; padding: 1rem 1.2rem;
    box-shadow: 0 2px 8px rgba(59,130,246,.15);
}
.kpi-value { font-size: 1.9rem; font-weight: 700; color: #1E3A8A; line-height:1.1; }
.kpi-label { font-size: .8rem; color: #6B7280; margin-top:.2rem; }
.kpi-delta { font-size: .75rem; color: #059669; font-weight: 600; }

.section-title {
    font-size: 1.15rem; font-weight: 700; color: #1E3A8A;
    border-bottom: 2px solid #BFDBFE; padding-bottom: .4rem; margin: 1.2rem 0 .8rem;
}
.chip {
    display: inline-block; background: #EFF6FF; color: #1D4ED8;
    border: 1px solid #BFDBFE; border-radius: 9999px;
    font-size: .72rem; padding: .1rem .55rem; margin: .15rem;
}
.badge-green  { background:#D1FAE5; color:#065F46; border-color:#6EE7B7; }
.badge-orange { background:#FEF3C7; color:#92400E; border-color:#FCD34D; }
.badge-red    { background:#FEE2E2; color:#991B1B; border-color:#FCA5A5; }

.stProgress > div > div > div > div { background-color: #3B82F6 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# 常量 & 预编译正则
# ─────────────────────────────────────────────────────────────────────────────
PERIOD_TIMES = {
    1:  ("08:00","08:50"),  2:  ("09:00","09:50"),
    3:  ("10:10","11:00"),  4:  ("11:10","12:00"),
    5:  ("13:30","14:20"),  6:  ("14:30","15:20"),
    7:  ("15:30","16:20"),  8:  ("16:30","17:20"),
    9:  ("17:30","18:20"),  10: ("18:30","19:20"),
    11: ("19:30","20:20"),  12: ("20:30","21:20"),
}
ALL_PERIODS  = list(range(1, 13))
TOTAL_WEEKS  = 17
DAY_MAP  = {"周一":0,"周二":1,"周三":2,"周四":3,"周五":4,"周六":5,"周日":6}
DAY_NAMES = ["周一","周二","周三","周四","周五","周六","周日"]

SLOT_RE = re.compile(r'(周[一二三四五六日])第([^节]+)节\{第([^}]+)周\}')
PSPLIT  = re.compile(r"[、，,]")
PRANGE  = re.compile(r"(\d+)\s*[-–]\s*(\d+)")
WSPLIT  = re.compile(r"[,，]")
WRANGE  = re.compile(r"(\d+)\s*[-–]\s*(\d+)")

CACHE_DIR = Path(tempfile.gettempdir()) / "schedule_cache"
CACHE_DIR.mkdir(exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# 工具函数 - 修复版
# ─────────────────────────────────────────────────────────────────────────────
def file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_periods(period_str: str) -> list:
    """解析节次字符串，支持逗号列表（7、8、9）和范围（7-9）两种格式"""
    periods = []
    # 先清理字符串
    period_str = period_str.strip()
    if not period_str:
        return []

    # 分割多个节次（支持中文顿号、英文逗号）
    for part in re.split(r'[、，,\s]+', period_str):
        part = part.strip()
        if not part:
            continue

        # 检查是否是范围（如 7-9 或 7–9）
        range_match = re.match(r'(\d+)\s*[-–]\s*(\d+)', part)
        if range_match:
            start = int(range_match.group(1))
            end = int(range_match.group(2))
            periods.extend(range(start, end + 1))
        else:
            try:
                periods.append(int(part))
            except ValueError:
                continue

    # 过滤有效节次范围
    return [p for p in periods if 1 <= p <= 12]


def parse_weeks(week_str: str) -> set:
    """解析周次字符串，返回周次集合"""
    weeks = set()
    week_str = week_str.strip()
    if not week_str:
        return weeks

    # 分割多个周次（支持中文顿号、英文逗号）
    for part in re.split(r'[、，,\s]+', week_str):
        part = part.strip()
        if not part:
            continue

        # 检查是否是范围（如 1-17 或 1–17）
        range_match = re.match(r'(\d+)\s*[-–]\s*(\d+)', part)
        if range_match:
            start = int(range_match.group(1))
            end = int(range_match.group(2))
            weeks.update(range(start, end + 1))
        else:
            try:
                weeks.add(int(part))
            except ValueError:
                continue

    # 限制在有效周次范围内
    return weeks & set(range(1, TOTAL_WEEKS + 1))



# ─────────────────────────────────────────────────────────────────────────────
# 逐行解析 XLSX（openpyxl 流模式，低内存）
# ─────────────────────────────────────────────────────────────────────────────
def parse_xlsx(xlsx_path: str) -> Tuple[Optional[np.ndarray], List[str]]:
    """解析单个 XLSX"""
    print(f"正在解析 XLSX 文件: {xlsx_path}")

    # 磁盘缓存检查
    cache_key = file_md5(xlsx_path)
    cache_path = CACHE_DIR / f"{cache_key}.pkl"
    if cache_path.exists():
        try:
            with open(cache_path, "rb") as f:
                return pickle.load(f)
        except Exception:
            cache_path.unlink(missing_ok=True)

    issues: List[str] = []
    occ = np.zeros((TOTAL_WEEKS, 7, 12), dtype=bool)
    col_idx: Optional[int] = None

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # 自动扫描前5行寻找"上课时间"列（兼容有/无合并标题行的 xlsx）
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(row)
            if len(all_rows) >= 5:
                # 一旦找到 header 就继续用 rows_iter 读剩余行
                break

        header_row_idx = None
        for ri, row in enumerate(all_rows):
            for ci, cell in enumerate(row):
                if cell is not None and "上课时间" in str(cell):
                    col_idx = ci
                    header_row_idx = ri
                    break
            if col_idx is not None:
                break

        if col_idx is None:
            wb.close()
            return None, [f"未找到'上课时间'列（已扫描前{len(all_rows)}行）"]

        unmatched = 0

        def _process_row(row):
            nonlocal unmatched
            if col_idx >= len(row):
                return
            cell = row[col_idx]
            if cell is None:
                return
            for part in str(cell).split(";"):
                part = part.strip()
                m = SLOT_RE.search(part)
                if not m:
                    if part and re.search(r'周[一二三四五六日]', part):
                        unmatched += 1
                    continue  # 继续处理下一个 part
                day_name = m.group(1)
                if day_name not in DAY_MAP:
                    continue
                day = DAY_MAP[day_name]
                periods = parse_periods(m.group(2))
                weeks = parse_weeks(m.group(3))
                for wk in weeks:
                    for p in periods:
                        if 1 <= p <= 12:
                            occ[wk - 1, day, p - 1] = True

        # 处理已缓冲的数据行
        for row in all_rows[header_row_idx + 1:]:
            _process_row(row)
        # 流式读取剩余行（节省内存）
        for row in ws.iter_rows(min_row=len(all_rows) + 1, values_only=True):
            _process_row(row)

        wb.close()

    except Exception as e:
        return None, [f"读取失败: {e}"]

    if unmatched:
        issues.append(f"有 {unmatched} 个时间片段格式不匹配")
    total_slots = int(occ.sum())
    if total_slots == 0:
        issues.append("课表节次全为零，可能格式异常")
    elif total_slots > 600:
        issues.append(f"节次总数偏多({total_slots})，请检查格式")

    result = (occ, issues)
    try:
        with open(cache_path, "wb") as f:
            pickle.dump(result, f)
    except Exception:
        pass

    return result


## ─────────────────────────────────────────────────────────────────────────────
# 解析 .xls（旧版 Excel，需要 xlrd >= 2.0.1）- 修复版
# ─────────────────────────────────────────────────────────────────────────────
def parse_xls(xls_path: str) -> Tuple[Optional[np.ndarray], List[str]]:
    """
    解析单个 .xls 文件，逻辑与 parse_xlsx 完全一致。
    需要安装 xlrd：pip install xlrd>=2.0.1
    """
    print(f"\n=== 进入 parse_xls 函数 ===")
    print(f"解析路径: {xls_path}")
    print(f"文件是否存在: {os.path.exists(xls_path)}")

    if not os.path.exists(xls_path):
        print(f"错误: 文件不存在!")
        return None, ["文件不存在"]

    if not _XLRD_OK:
        print("xlrd 未安装")
        return None, ["缺少依赖：请运行 `pip install xlrd>=2.0.1` 后重试"]


    # ── 磁盘缓存检查 ──
    cache_key = file_md5(xls_path)
    cache_path = CACHE_DIR / f"{cache_key}.pkl"
    if cache_path.exists():
        try:
            with open(cache_path, "rb") as f:
                return pickle.load(f)
        except Exception:
            cache_path.unlink(missing_ok=True)

    issues: List[str] = []
    occ = np.zeros((TOTAL_WEEKS, 7, 12), dtype=bool)
    col_idx: Optional[int] = None

    try:
        # 打印调试信息
        print(f"正在解析 XLS 文件: {xls_path}")

        wb = xlrd.open_workbook(xls_path)
        ws = wb.sheet_by_index(0)

        print(f"工作表名称: {ws.name}, 行数: {ws.nrows}, 列数: {ws.ncols}")

        if ws.nrows < 1:
            return None, ["文件为空"]

        # 打印前5行内容看看数据结构
        print("\n=== 前5行内容 ===")
        for i in range(min(5, ws.nrows)):
            row = ws.row_values(i)
            print(f"行 {i + 1}: {row}")

        # 自动扫描前10行寻找"上课时间"列
        header_row_idx = None
        for try_row in range(min(10, ws.nrows)):
            row_values = ws.row_values(try_row)
            for c, cell_val in enumerate(row_values):
                cell_str = str(cell_val) if cell_val else ""
                print(f"  检查行 {try_row + 1}, 列 {c + 1}: '{cell_str}'")
                if cell_val and "上课时间" in cell_str:
                    col_idx = c
                    header_row_idx = try_row
                    print(f"✓ 找到'上课时间'列在第 {try_row + 1} 行, 第 {c + 1} 列, 值: '{cell_val}'")
                    break
            if col_idx is not None:
                break

        if col_idx is None:
            print("❌ 未找到'上课时间'列")
            return None, [f"未找到'上课时间'列（已扫描前{min(10, ws.nrows)}行）"]

        # 从标题行下一行开始解析
        data_start_row = header_row_idx + 1
        unmatched = 0
        parsed_count = 0

        for r in range(data_start_row, ws.nrows):
            row_values = ws.row_values(r)
            if col_idx >= len(row_values):
                continue

            cell_val = row_values[col_idx]
            if not cell_val or str(cell_val).strip() == "":
                continue

            # 处理单元格内容
            cell_str = str(cell_val).strip()
            print(f"行 {r + 1}: {cell_str[:50]}...")  # 打印前50个字符调试

            # 分割多个课程（如果有分号）
            for part in cell_str.split(";"):
                part = part.strip()
                if not part:
                    continue

                # 使用更宽松的正则表达式
                m = re.search(r'(周[一二三四五六日])第([^节]+)节\{第([^}]+)周\}', part)
                if not m:
                    m = re.search(r'(周[一二三四五六日])第([^\{]+)\{第([^\}]+)\}', part)

                if not m:
                    if part and re.search(r'周[一二三四五六日]', part):
                        unmatched += 1
                        print(f"  无法解析: {part}")
                    continue

                day_name = m.group(1)
                if day_name not in DAY_MAP:
                    print(f"  未知星期: {day_name}")
                    continue

                day = DAY_MAP[day_name]
                period_str = m.group(2).strip()
                week_str = m.group(3).strip()

                periods = parse_periods(period_str)
                if not periods:
                    print(f"  无法解析节次: {period_str}")
                    continue

                weeks = parse_weeks(week_str)
                if not weeks:
                    print(f"  无法解析周次: {week_str}")
                    continue

                for wk in weeks:
                    for p in periods:
                        if 1 <= p <= 12:
                            occ[wk - 1, day, p - 1] = True
                            parsed_count += 1

        print(f"解析完成: 成功解析 {parsed_count} 个课程时段, 无法解析 {unmatched} 个片段")

    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, [f"读取失败: {str(e)}"]

    if unmatched:
        issues.append(f"有 {unmatched} 个时间片段格式不匹配")

    total_slots = int(occ.sum())
    if total_slots == 0:
        issues.append("课表节次全为零，可能格式异常")
    elif total_slots > 600:
        issues.append(f"节次总数偏多({total_slots})，请检查格式")

    result = (occ, issues)
    try:
        with open(cache_path, "wb") as f:
            pickle.dump(result, f)
    except Exception:
        pass

    return result

def parse_file(file_path: str) -> Tuple[Optional[np.ndarray], List[str]]:
    """根据扩展名自动分发到对应解析器"""
    if file_path.lower().endswith(".xls"):
        return parse_xls(file_path)
    return parse_xlsx(file_path)


# ─────────────────────────────────────────────────────────────────────────────
# 异步并行解析（asyncio + ThreadPoolExecutor）
# ─────────────────────────────────────────────────────────────────────────────
async def _parse_one_async(executor, name: str, path: str):
    loop = asyncio.get_event_loop()
    occ, issues = await loop.run_in_executor(executor, parse_file, path)
    return name, occ, issues


async def parse_all_async(paths: List[Tuple[str, str]], max_workers: int = 4,
                          progress_cb=None) -> Tuple[Dict, Dict]:
    """并行解析所有文件，返回 (all_occ, all_issues)"""
    all_occ: Dict[str, np.ndarray] = {}
    all_issues: Dict[str, List[str]] = {}
    done = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        tasks = [_parse_one_async(executor, name, path) for name, path in paths]
        for coro in asyncio.as_completed(tasks):
            name, occ, issues = await coro
            done += 1
            if progress_cb:
                progress_cb(done / len(paths), f"已解析 {done}/{len(paths)}: {name}")
            if occ is not None:
                all_occ[name] = occ
            if issues:
                all_issues[name] = issues

    return all_occ, all_issues


def run_parse(paths, max_workers, progress_cb=None):
    """同步入口（Streamlit 调用）"""
    return asyncio.run(parse_all_async(paths, max_workers, progress_cb))


# ─────────────────────────────────────────────────────────────────────────────
# 稀疏矩阵 occ_sum 聚合
# ─────────────────────────────────────────────────────────────────────────────
def build_free4d(all_occ: Dict[str, np.ndarray]) -> Tuple[np.ndarray, int]:
    """
    返回 free4d[TOTAL_WEEKS, 7, 12] 和人数 n。
    内部使用 scipy COO 稀疏矩阵累加占用，节省内存。
    """
    n = len(all_occ)
    shape_flat = TOTAL_WEEKS * 7 * 12
    row_list, col_list, data_list = [], [], []

    for idx, occ in enumerate(all_occ.values()):
        coords = np.argwhere(occ)           # (K, 3)
        flat   = coords[:, 0] * 84 + coords[:, 1] * 12 + coords[:, 2]
        row_list.append(flat)
        col_list.append(np.full(len(flat), idx, dtype=np.int32))
        data_list.append(np.ones(len(flat), dtype=np.int8))

    if row_list:
        rows = np.concatenate(row_list)
        cols = np.concatenate(col_list)
        data = np.concatenate(data_list)
        sp = sparse.coo_matrix((data, (rows, cols)), shape=(shape_flat, n))
        occ_sum = np.array(sp.sum(axis=1)).reshape(TOTAL_WEEKS, 7, 12)
    else:
        occ_sum = np.zeros((TOTAL_WEEKS, 7, 12), dtype=int)

    free4d = n - occ_sum
    return free4d.astype(int), n


# ─────────────────────────────────────────────────────────────────────────────
# ── 可视化函数（Plotly，中文标签）────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
_COLORSCALE_MAIN = "YlGnBu"

def _layout_defaults(fig, title="", h=500):
    fig.update_layout(
        title=dict(text=title, font=dict(size=16, color="#1E3A8A", family="Inter, sans-serif")),
        height=h,
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", color="#374151"),
        margin=dict(t=60, b=40, l=40, r=30),
    )
    return fig


def chart_heatmap(free4d: np.ndarray, n: int, week: Optional[int] = None, english: bool = False):
    """周度热力图"""
    if english:
        day_names_en = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        period_labels = [f"Period {p} ({PERIOD_TIMES[p][0]})" for p in ALL_PERIODS]
        title = f"Week {week + 1} Free Period Heatmap (Total: {n} People)"
    else:
        day_names_en = DAY_NAMES
        period_labels = [f"第{p}节 {PERIOD_TIMES[p][0]}" for p in ALL_PERIODS]
        title = f"第{week + 1}周 空余人数热力图（共{n}人）"

    if week is not None:
        z = free4d[week]
    else:
        z = free4d.mean(axis=0)
        title = f"17周平均空余人数热力图（共{n}人）" if not english else f"17-Week Average Free Period Heatmap (Total: {n} People)"

    fig = go.Figure(go.Heatmap(
        z=z.T, x=day_names_en,
        y=period_labels,
        colorscale="YlGnBu", zmin=0, zmax=n,
        text=np.round(z.T, 1), texttemplate="%{text}",
        textfont={"size": 10},
        hovertemplate="<b>%{x}</b><br>%{y}<br>Free: %{z:.1f} people<extra></extra>",
    ))
    fig.update_yaxes(autorange="reversed", title="Period" if english else "节次")
    fig.update_xaxes(title="Day" if english else "星期")
    return _layout_defaults(fig, title, h=580)

def chart_weekly_trend(free4d: np.ndarray, n: int):
    """各周空闲率趋势（柱状 + 折线）"""
    max_pos = n * 7 * 12
    pct = free4d.sum(axis=(1, 2)) / max_pos * 100
    weeks = list(range(1, TOTAL_WEEKS + 1))
    colors = ["#10B981" if p >= 60 else "#F59E0B" if p >= 40 else "#EF4444" for p in pct]

    fig = make_subplots(specs=[[{"secondary_y": False}]])
    fig.add_trace(go.Bar(
        x=weeks, y=pct, marker_color=colors, name="空闲率",
        text=[f"{p:.1f}%" for p in pct], textposition="outside",
        hovertemplate="第%{x}周<br>空闲率: %{y:.1f}%<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=weeks, y=pct, mode="lines+markers", name="趋势线",
        line=dict(color="#1D4ED8", width=2.5),
        marker=dict(size=7, color="#1D4ED8"),
    ))
    fig.add_hrect(y0=60, y1=105, fillcolor="#D1FAE5", opacity=0.2, line_width=0)
    fig.add_hrect(y0=40, y1=60,  fillcolor="#FEF3C7", opacity=0.2, line_width=0)
    fig.add_hrect(y0=0,  y1=40,  fillcolor="#FEE2E2", opacity=0.2, line_width=0)
    fig.add_hline(y=60, line_dash="dash", line_color="#059669",
                  annotation_text="优 60%", annotation_position="bottom right")
    fig.add_hline(y=40, line_dash="dash", line_color="#D97706",
                  annotation_text="良 40%", annotation_position="bottom right")
    fig.update_xaxes(
        tickmode="array", tickvals=weeks, ticktext=[f"第{w}周" for w in weeks], title="周次"
    )
    fig.update_yaxes(range=[0, 110], title="空闲率 (%)")
    return _layout_defaults(fig, f"各周整体空闲率趋势（共{n}人）", h=480)


def chart_heatmap_with_names(free4d: np.ndarray, n: int, all_occ: Dict[str, np.ndarray], week: int,
                             threshold: float = 5.0):
    """
    生成带人名的周度热力图
    - 绿色格子：全员空闲
    - 红色格子：低空闲率（标注缺席人员）
    - 黄色格子：正常空闲
    """
    # 计算每个时段的有课人数
    occ_sum = np.zeros((7, 12), dtype=int)
    for occ in all_occ.values():
        occ_sum += occ[week].astype(int)

    # 找出低空闲率时段（缺席人数 <= threshold%）
    absent_count = n - free4d[week]  # 每个时段缺席人数
    absent_count = n - free4d[week]
    low_free_mask = absent_count < (n * threshold / 100)  # 缺席人数少于阈值百分比

    # 为每个时段收集缺席人员名单
    absent_names = np.empty((7, 12), dtype=object)
    for d in range(7):
        for p in range(12):
            if low_free_mask[d, p]:
                names = []
                for name, occ in all_occ.items():
                    if occ[week, d, p]:  # 该人有课
                        names.append(name)
                absent_names[d, p] = "\n".join(names[:3]) + ("..." if len(names) > 3 else "")
            else:
                absent_names[d, p] = ""

    # 创建热力图
    fig = go.Figure()

    # 基础热力图（背景色）
    fig.add_trace(go.Heatmap(
        z=free4d[week].T,
        x=DAY_NAMES,
        y=[f"第{p}节 {PERIOD_TIMES[p][0]}" for p in ALL_PERIODS],
        colorscale=[
            [0, '#ef4444'],  # 红色：低空闲率
            [threshold / 100, '#f59e0b'],  # 橙色：中等空闲
            [0.5, '#f59e0b'],
            [1, '#10b981']  # 绿色：高空闲率
        ],
        zmin=0,
        zmax=n,
        showscale=True,
        colorbar=dict(title="空余人数"),
        hovertemplate="<b>%{x}</b><br>%{y}<br>空余: %{z}人<br>缺席: %{customdata[0]}人<br>缺席人员: %{customdata[1]}<extra></extra>",
        customdata=np.stack([absent_count.T, absent_names.T], axis=-1),
        text=np.round(free4d[week].T, 1),
        texttemplate="%{text}",
        textfont={"size": 10},
    ))

    # 添加全员空闲标记
    fully_free_mask = free4d[week] == n
    if fully_free_mask.any():
        # 在全员空闲的格子上添加星星标记
        for d in range(7):
            for p in range(12):
                if fully_free_mask[d, p]:
                    fig.add_annotation(
                        x=DAY_NAMES[d],
                        y=f"第{p + 1}节 {PERIOD_TIMES[p + 1][0]}",
                        text="⭐",
                        showarrow=False,
                        font=dict(size=16),
                    )

    fig.update_layout(
        title=dict(
            text=f"第{week + 1}周 空余人数热力图（含缺席标注）",
            font=dict(size=16, color="#1E3A8A")
        ),
        height=600,
        xaxis_title="星期",
        yaxis_title="节次",
        yaxis=dict(autorange="reversed"),
    )

    return fig


def chart_low_free_persons(free4d: np.ndarray, n: int, all_occ: Dict[str, np.ndarray], threshold: float = 5.0):
    """
    生成低空闲率人员分布图
    显示每个时段缺席的人员名单
    """
    # 计算每个人员的总空闲率
    total_slots = TOTAL_WEEKS * 7 * 12
    person_free_rate = {}
    for name, occ in all_occ.items():
        busy = int(occ.sum())
        free_rate = (total_slots - busy) / total_slots * 100
        person_free_rate[name] = free_rate

    # 找出低空闲率人员
    low_free_persons = [name for name, rate in person_free_rate.items() if rate < threshold]

    if not low_free_persons:
        return None

    # 创建矩阵显示低空闲率人员的有课时段
    presence_matrix = np.zeros((len(low_free_persons), TOTAL_WEEKS * 7 * 12), dtype=bool)
    for i, name in enumerate(low_free_persons):
        occ = all_occ[name]
        presence_matrix[i] = occ.flatten()

    # 创建热力图
    fig = go.Figure()

    # 周次标签
    week_labels = []
    for w in range(TOTAL_WEEKS):
        for d in range(7):
            for p in range(12):
                week_labels.append(f"W{w + 1}-{DAY_NAMES[d]}-P{p + 1}")

    fig.add_trace(go.Heatmap(
        z=presence_matrix,
        x=list(range(TOTAL_WEEKS * 7 * 12)),
        y=low_free_persons,
        colorscale=[[0, '#10b981'], [1, '#ef4444']],
        showscale=False,
        hovertemplate="人员: %{y}<br>时段: %{customdata}<br>状态: %{z}<extra></extra>",
        customdata=[week_labels] * len(low_free_persons),
    ))

    # 添加周次分隔线
    shapes = []
    for w in range(1, TOTAL_WEEKS):
        shapes.append(dict(
            type="line",
            x0=w * 7 * 12,
            y0=-0.5,
            x1=w * 7 * 12,
            y1=len(low_free_persons) - 0.5,
            line=dict(color="white", width=2, dash="dash"),
        ))

    fig.update_layout(
        title=dict(
            text=f"低空闲率人员 ({threshold}%) 有课时段分布",
            font=dict(size=16, color="#1E3A8A")
        ),
        height=max(400, len(low_free_persons) * 30),
        xaxis_title="时段 (按周次分隔)",
        yaxis_title="人员",
        shapes=shapes,
    )

    # 添加周次标注
    for w in range(TOTAL_WEEKS):
        fig.add_annotation(
            x=w * 7 * 12 + 3.5 * 7 * 12,
            y=len(low_free_persons),
            text=f"第{w + 1}周",
            showarrow=False,
            font=dict(size=10, color="white"),
            yanchor="bottom",
        )

    return fig

def chart_radar(free4d: np.ndarray, week: int):
    """雷达图：各天空闲节次"""
    daily = free4d[week].sum(axis=1)
    max_p = 12
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=list(daily) + [daily[0]],
        theta=DAY_NAMES + [DAY_NAMES[0]],
        fill="toself", name=f"第{week+1}周",
        line=dict(color="#2563EB", width=2.5),
        fillcolor="rgba(37,99,235,.25)",
    ))
    fig.add_trace(go.Scatterpolar(
        r=[max_p] * 8, theta=DAY_NAMES + [DAY_NAMES[0]],
        mode="lines", line=dict(color="#9CA3AF", dash="dot"),
        name="最大可能", showlegend=True,
    ))
    fig.update_layout(
        polar=dict(radialaxis=dict(range=[0, max_p], tickfont=dict(size=9))),
        showlegend=True, height=440,
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        title=dict(text=f"第{week+1}周 各天空闲节次分布", font=dict(size=15, color="#1E3A8A")),
        margin=dict(t=60, b=20, l=40, r=40),
    )
    return fig


def chart_fully_free_heatmap(free4d: np.ndarray, n: int):
    """全员空闲周数分布热力图"""
    fw = np.array([[int((free4d[:, d, p] == n).sum()) for p in range(12)] for d in range(7)])
    fig = go.Figure(go.Heatmap(
        z=fw.T, x=DAY_NAMES,
        y=[f"第{p}节 {PERIOD_TIMES[p][0]}" for p in ALL_PERIODS],
        colorscale="Reds", zmin=0, zmax=TOTAL_WEEKS,
        text=fw.T, texttemplate="%{text}",
        textfont={"size": 10},
        hovertemplate="<b>%{x}</b><br>%{y}<br>全员空闲: %{z} 周<extra></extra>",
    ))
    fig.update_yaxes(autorange="reversed", title="节次")
    fig.update_xaxes(title="星期")
    return _layout_defaults(fig, f"各时段全员空闲周数（共{TOTAL_WEEKS}周）", h=580)


def display_person_conflict_matrix(all_occ: Dict[str, np.ndarray], week: int, period: int, day: int):
    """
    显示特定时段的有课人员
    """
    persons_with_class = []
    persons_free = []

    for name, occ in all_occ.items():
        if occ[week, day, period]:
            persons_with_class.append(name)
        else:
            persons_free.append(name)

    return persons_with_class, persons_free

def chart_3d_surface(free4d: np.ndarray, n: int, week: Optional[int] = None):
    """3D 曲面图"""
    Z = free4d[week].T if week is not None else free4d.mean(axis=0).T
    title = f"第{week+1}周 3D空闲分布" if week is not None else "17周平均 3D空闲分布"
    fig = go.Figure(go.Surface(
        z=Z, x=DAY_NAMES, y=[f"第{p}节" for p in ALL_PERIODS],
        colorscale="Viridis",
        hovertemplate="星期: %{x}<br>节次: %{y}<br>空闲: %{z:.1f}人<extra></extra>",
    ))
    fig.update_layout(
        scene=dict(
            xaxis_title="星期", yaxis_title="节次", zaxis_title="空闲人数",
            camera=dict(eye=dict(x=1.5, y=1.5, z=1.5)),
            bgcolor="rgba(0,0,0,0)",
        ),
        height=600, paper_bgcolor="rgba(0,0,0,0)",
        title=dict(text=title, font=dict(size=15, color="#1E3A8A")),
        margin=dict(t=60, b=20),
    )
    return fig


def chart_bubble(free4d: np.ndarray, n: int):
    """气泡图：每个时间格的空闲率（横轴=星期，纵轴=节次，气泡大小=空闲率）"""
    avg = free4d.mean(axis=0)  # (7,12)
    rows = []
    for d in range(7):
        for p in range(12):
            pct = avg[d, p] / n * 100
            rows.append(dict(
                星期=DAY_NAMES[d],
                节次=f"第{p+1}节",
                节次序号=p+1,
                空闲率=pct,
                空闲人数=avg[d, p],
                颜色=("#10B981" if pct >= 70 else "#F59E0B" if pct >= 40 else "#EF4444"),
            ))
    df = pd.DataFrame(rows)
    fig = px.scatter(
        df, x="星期", y="节次序号", size="空闲率",
        color="空闲率", color_continuous_scale="RdYlGn",
        range_color=[0, 100],
        hover_data={"星期": True, "节次": True, "空闲率": ":.1f", "空闲人数": ":.1f", "节次序号": False},
        labels={"节次序号": "节次", "空闲率": "空闲率(%)"},
        size_max=30,
        category_orders={"星期": DAY_NAMES},
    )
    fig.update_yaxes(
        tickmode="array",
        tickvals=list(range(1, 13)),
        ticktext=[f"第{p}节" for p in range(1, 13)],
        autorange="reversed", title="节次",
    )
    fig.update_xaxes(title="星期")
    return _layout_defaults(fig, "时间格空闲率气泡图（气泡越大=越空闲）", h=560)


def chart_person_busyload(all_occ: Dict[str, np.ndarray]):
    """各人有课总节次分布（水平柱状图 + 箱线图）"""
    data = sorted([(name, int(occ.sum())) for name, occ in all_occ.items()], key=lambda x: x[1])
    names, vals = zip(*data) if data else ([], [])

    fig = make_subplots(
        rows=1, cols=2,
        column_widths=[0.68, 0.32],
        subplot_titles=("各人有课总节次排名", "节次分布"),
    )

    colors = ["#EF4444" if v > np.percentile(vals, 80)
              else "#F59E0B" if v > np.percentile(vals, 50)
              else "#10B981" for v in vals]

    fig.add_trace(go.Bar(
        y=list(names), x=list(vals), orientation="h",
        marker_color=colors, name="有课节次",
        hovertemplate="%{y}: %{x} 节<extra></extra>",
    ), row=1, col=1)

    fig.add_trace(go.Box(
        y=list(vals), name="分布",
        marker_color="#3B82F6", boxmean="sd",
        hovertemplate="节次: %{y}<extra></extra>",
    ), row=1, col=2)

    fig.update_yaxes(title="姓名/班级", row=1, col=1)
    fig.update_xaxes(title="有课节次", row=1, col=1)
    fig.update_yaxes(title="节次数", row=1, col=2)

    h = max(400, len(names) * 22 + 100)
    return _layout_defaults(fig, "各人课业负担分析", h=h)


def chart_grouped_bar(free4d: np.ndarray, n: int):
    """分组柱状图：不同节次段（上午/下午/晚上）各天平均空闲率"""
    avg = free4d.mean(axis=0)  # (7,12)
    groups = {"上午(1-4节)": list(range(0,4)), "下午(5-9节)": list(range(4,9)), "晚上(10-12节)": list(range(9,12))}

    fig = go.Figure()
    palette = ["#3B82F6", "#10B981", "#F59E0B"]
    for (label, idxs), color in zip(groups.items(), palette):
        pct = [avg[d, idxs].mean() / n * 100 for d in range(7)]
        fig.add_trace(go.Bar(
            name=label, x=DAY_NAMES, y=pct,
            marker_color=color,
            text=[f"{v:.0f}%" for v in pct], textposition="outside",
            hovertemplate=f"<b>{label}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>",
        ))
    fig.update_layout(
        barmode="group", yaxis=dict(range=[0, 115], title="平均空闲率 (%)"),
        xaxis_title="星期",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    return _layout_defaults(fig, "各天各时段平均空闲率（分组）", h=460)


# ─────────────────────────────────────────────────────────────────────────────
# Matplotlib 图表（用于 PNG/PDF 导出，中文字体已配置）
# ─────────────────────────────────────────────────────────────────────────────
def mpl_heatmap_bytes(free4d: np.ndarray, n: int, week: Optional[int] = None,
                      fmt: str = "png") -> bytes:
    """生成 matplotlib 热力图并返回字节流（PNG 或 PDF）"""
    z = free4d[week].T if week is not None else free4d.mean(axis=0).T
    title = (f"第{week+1}周 空余热力图" if week is not None else "17周平均空余热力图")

    fig, ax = plt.subplots(figsize=(10, 7))
    im = ax.imshow(z, cmap="YlGnBu", vmin=0, vmax=n, aspect="auto")
    plt.colorbar(im, ax=ax, fraction=0.03, pad=0.02, label="空余人数")
    ax.set_xticks(range(7));  ax.set_xticklabels(DAY_NAMES, fontsize=11)
    ax.set_yticks(range(12)); ax.set_yticklabels([f"第{p}节\n{PERIOD_TIMES[p][0]}" for p in ALL_PERIODS], fontsize=8)
    ax.set_title(f"{title}（共{n}人）", fontsize=14, fontweight="bold", pad=12)
    ax.set_xlabel("星期"); ax.set_ylabel("节次")
    for (r, c), val in np.ndenumerate(z):
        ax.text(c, r, f"{val:.0f}", ha="center", va="center", fontsize=9,
                color="white" if val < n*0.55 else "#1a1a1a")
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf.read()


def mpl_trend_bytes(free4d: np.ndarray, n: int, fmt: str = "png") -> bytes:
    max_pos = n * 7 * 12
    pct = free4d.sum(axis=(1, 2)) / max_pos * 100
    weeks = list(range(1, TOTAL_WEEKS + 1))
    colors = ["#10B981" if p >= 60 else "#F59E0B" if p >= 40 else "#EF4444" for p in pct]

    fig, ax = plt.subplots(figsize=(12, 5))
    bars = ax.bar(weeks, pct, color=colors, alpha=0.85, zorder=2)
    ax.plot(weeks, pct, "o-", color="#1D4ED8", linewidth=2, markersize=6, zorder=3)
    ax.axhline(60, ls="--", color="#059669", lw=1.2, label="60% 基准")
    ax.axhline(40, ls="--", color="#D97706", lw=1.2, label="40% 基准")
    ax.set_xticks(weeks); ax.set_xticklabels([f"第{w}周" for w in weeks], rotation=45, fontsize=8)
    ax.set_ylabel("空闲率 (%)"); ax.set_ylim(0, 110)
    ax.set_title(f"各周整体空闲率趋势（共{n}人）", fontsize=13, fontweight="bold")
    ax.grid(axis="y", ls=":", alpha=0.5, zorder=1)
    ax.legend(fontsize=9)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Excel 多 Sheet 导出
# ─────────────────────────────────────────────────────────────────────────────
def build_excel_export(free4d: np.ndarray, n: int,
                       all_occ: Dict[str, np.ndarray]) -> bytes:
    """生成含多个 Sheet 的 Excel 文件"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        wb  = writer.book
        hdr = wb.add_format({"bold": True, "bg_color": "#1E3A8A",
                              "font_color": "#FFFFFF", "border": 1, "align": "center"})
        num = wb.add_format({"border": 1, "align": "center", "num_format": "0.0"})
        pct_fmt = wb.add_format({"border": 1, "align": "center", "num_format": "0.0%"})
        green = wb.add_format({"border":1,"align":"center","bg_color":"#D1FAE5","num_format":"0"})
        red   = wb.add_format({"border":1,"align":"center","bg_color":"#FEE2E2","num_format":"0"})

        # ── Sheet 1: 17周平均空闲率 ──
        avg = free4d.mean(axis=0)
        df1 = pd.DataFrame(avg, index=DAY_NAMES,
                           columns=[f"第{p}节\n{PERIOD_TIMES[p][0]}" for p in ALL_PERIODS])
        df1.to_excel(writer, sheet_name="17周平均空闲人数", index=True)
        ws1 = writer.sheets["17周平均空闲人数"]
        ws1.set_column("A:A", 8)
        ws1.set_column("B:M", 10)

        # ── Sheet 2: 各周空闲率 ──
        max_pos = n * 7 * 12
        rows2 = []
        for w in range(TOTAL_WEEKS):
            free_sum = int(free4d[w].sum())
            rows2.append({
                "周次": f"第{w+1}周",
                "空闲节次总数": free_sum,
                "空闲率": free_sum / max_pos,
                "全员空闲时段数": int((free4d[w] == n).sum()),
            })
        pd.DataFrame(rows2).to_excel(writer, sheet_name="各周空闲率", index=False)

        # ── Sheet 3: 全员空闲明细 ──
        rows3 = []
        for w in range(TOTAL_WEEKS):
            for d in range(7):
                for p in range(12):
                    if free4d[w, d, p] == n:
                        rows3.append({
                            "周次": f"第{w+1}周", "星期": DAY_NAMES[d],
                            "节次": p+1,
                            "时间段": f"{PERIOD_TIMES[p+1][0]}–{PERIOD_TIMES[p+1][1]}",
                        })
        pd.DataFrame(rows3).to_excel(writer, sheet_name="全员空闲明细", index=False)

        # ── Sheet 4: 各人课业负担 ──
        rows4 = []
        for name, occ in all_occ.items():
            total_busy = int(occ.sum())
            rows4.append({
                "姓名/班级": name,
                "有课总节次": total_busy,
                "空闲总节次": TOTAL_WEEKS * 7 * 12 - total_busy,
                "空闲率": 1 - total_busy / (TOTAL_WEEKS * 7 * 12),
            })
        rows4.sort(key=lambda r: r["有课总节次"])
        pd.DataFrame(rows4).to_excel(writer, sheet_name="各人课业负担", index=False)

        # ── Sheet 5: 每周每天每节次详情 ──
        rows5 = []
        for w in range(TOTAL_WEEKS):
            for d in range(7):
                for p in range(12):
                    rows5.append({
                        "周次": w+1, "星期": DAY_NAMES[d], "节次": p+1,
                        "时间段": f"{PERIOD_TIMES[p+1][0]}–{PERIOD_TIMES[p+1][1]}",
                        "空余人数": int(free4d[w, d, p]),
                        "空余率": int(free4d[w, d, p]) / n if n > 0 else 0,
                        "全员空闲": "✓" if free4d[w, d, p] == n else "",
                    })
        pd.DataFrame(rows5).to_excel(writer, sheet_name="全量明细", index=False)

    buf.seek(0)
    return buf.read()


def build_csv_zip(free4d: np.ndarray, n: int,
                  all_occ: Dict[str, np.ndarray]) -> bytes:
    """打包多个 CSV 文件为 ZIP"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 17周平均
        avg = free4d.mean(axis=0)
        df = pd.DataFrame(avg, index=DAY_NAMES,
                          columns=[f"第{p}节" for p in ALL_PERIODS])
        zf.writestr("17周平均空闲人数.csv", df.to_csv(encoding="utf-8-sig"))

        # 全量明细
        rows = []
        for w in range(TOTAL_WEEKS):
            for d in range(7):
                for p in range(12):
                    rows.append({
                        "周次": w+1, "星期": DAY_NAMES[d], "节次": p+1,
                        "空余人数": int(free4d[w,d,p]),
                        "空余率%": round(free4d[w,d,p]/n*100, 1) if n else 0,
                        "全员空闲": free4d[w,d,p]==n,
                    })
        zf.writestr("全量明细.csv", pd.DataFrame(rows).to_csv(index=False, encoding="utf-8-sig"))

        # 各人负担
        rows2 = [{"姓名": k, "有课节次": int(v.sum())} for k, v in all_occ.items()]
        zf.writestr("各人课业负担.csv", pd.DataFrame(rows2).to_csv(index=False, encoding="utf-8-sig"))

    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# 报告生成
# ─────────────────────────────────────────────────────────────────────────────
def generate_report(free4d: np.ndarray, n: int,
                    all_occ: Dict[str, np.ndarray],
                    issues: Dict[str, List[str]]) -> str:
    lines = [
        "# 📅 班级课表空余时间统计报告",
        "",
        f"| 项目 | 数值 |",
        f"|------|------|",
        f"| 参与人数 | **{n} 人** |",
        f"| 统计周次 | 第 1–{TOTAL_WEEKS} 周 |",
        f"| 时间范围 | 08:00–21:20（含午休段标记） |",
        "",
    ]
    if issues:
        lines += ["## ⚠️ 数据问题", ""]
        for nm, iss_list in issues.items():
            for iss in iss_list:
                lines.append(f"- **{nm}**: {iss}")
        lines.append("")

    fw = np.array([[int((free4d[:, d, p] == n).sum()) for p in range(12)] for d in range(7)])
    cands = sorted([(fw[d, p], d, p+1) for d in range(7) for p in range(12)], reverse=True)
    lines += ["## 🏆 全员空闲时段 TOP 10", ""]
    lines.append("| 排名 | 时段 | 时间 | 全员空闲周数 |")
    lines.append("|------|------|------|------------|")
    for rank, (wks, d, p) in enumerate(cands[:10], 1):
        t = PERIOD_TIMES[p]
        lines.append(f"| {rank} | {DAY_NAMES[d]} 第{p}节 | {t[0]}–{t[1]} | **{wks}/{TOTAL_WEEKS}** 周 |")
    lines.append("")

    max_pos = n * 7 * 12
    lines += ["## 📈 各周空闲率", ""]
    lines.append("| 周次 | 空闲率 | 进度 |")
    lines.append("|------|--------|------|")
    for w in range(TOTAL_WEEKS):
        pct = free4d[w].sum() / max_pos * 100
        bar_len = int(pct / 5)
        bar = "█" * bar_len + "░" * (20 - bar_len)
        lines.append(f"| 第{w+1}周 | {pct:.1f}% | `{bar}` |")
    lines.append("")

    lines += ["## 🕐 各周全员共同空闲时段", ""]
    for w in range(TOTAL_WEEKS):
        slots = [f"{DAY_NAMES[d]}第{p+1}节"
                 for d in range(7) for p in range(12) if free4d[w, d, p] == n]
        if slots:
            show = "、".join(slots[:6]) + (f" …共{len(slots)}个" if len(slots) > 6 else "")
            lines.append(f"- 第{w+1}周：{show}")
        else:
            lines.append(f"- 第{w+1}周：*无全员共同空闲时段*")
    lines.append("")

    rows_p = sorted([(nm, int(occ.sum())) for nm, occ in all_occ.items()], key=lambda x: x[1])
    lines += ["## 👥 各人有课节次（升序）", ""]
    lines.append("| 姓名/班级 | 有课节次 |")
    lines.append("|-----------|---------|")
    for nm, busy in rows_p:
        lines.append(f"| {nm} | **{busy}** |")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# 输入处理：ZIP 解压 + 文件夹（多文件直传）
# ─────────────────────────────────────────────────────────────────────────────
def collect_xlsx_from_zip(uploaded_zip, add_debug) -> Tuple[str, List[Tuple[str, str]]]:
    """解压 ZIP，返回 (temp_dir, [(name, path)])"""
    add_debug("=" * 50)
    add_debug("【ZIP文件解压调试信息】")
    add_debug(f"ZIP文件大小: {len(uploaded_zip.getvalue())} 字节")

    tmp = tempfile.mkdtemp()
    zip_tmp = os.path.join(tmp, "upload.zip")
    add_debug(f"创建临时目录: {tmp}")
    add_debug(f"保存ZIP到: {zip_tmp}")

    with open(zip_tmp, "wb") as f:
        f.write(uploaded_zip.getvalue())
    add_debug(f"ZIP文件已保存")

    # 列出ZIP中的文件
    with zipfile.ZipFile(zip_tmp, "r") as zf:
        file_list = zf.namelist()
        add_debug(f"\nZIP中包含 {len(file_list)} 个文件/文件夹:")
        for i, name in enumerate(file_list[:10]):
            add_debug(f"  {i + 1}. {name}")
        if len(file_list) > 10:
            add_debug(f"  ... 还有 {len(file_list) - 10} 个文件")

        # 解压文件
        for member in zf.infolist():
            try:
                member.filename = member.filename.encode("cp437").decode("gbk")
            except Exception:
                pass
            try:
                zf.extract(member, tmp)
                add_debug(f"解压: {member.filename}")
            except Exception as e:
                add_debug(f"解压失败 {member.filename}: {e}")

    # 查找所有Excel文件
    add_debug("\n扫描解压后的文件:")
    paths = []
    for p in Path(tmp).rglob("*"):
        if p.suffix.lower() in (".xlsx", ".xls") and not p.name.startswith("~$"):
            add_debug(f"找到Excel文件: {p.name}")
            add_debug(f"  大小: {p.stat().st_size} 字节")
            paths.append((p.stem, str(p)))

    add_debug(f"\n总共找到 {len(paths)} 个Excel文件")
    add_debug("=" * 50)
    return tmp, paths

def collect_xlsx_from_files(uploaded_files, add_debug) -> Tuple[str, List[Tuple[str, str]]]:
    """将多上传文件保存到临时目录，返回 (temp_dir, [(name, path)])"""
    add_debug("=" * 50)
    add_debug("【多文件上传调试信息】")
    add_debug(f"收到 {len(uploaded_files)} 个文件")

    tmp = tempfile.mkdtemp()
    add_debug(f"创建临时目录: {tmp}")

    paths = []
    for uf in uploaded_files:
        dest = os.path.join(tmp, uf.name)
        add_debug(f"\n保存文件: {uf.name} -> {dest}")

        with open(dest, "wb") as f:
            f.write(uf.getvalue())
        add_debug(f"文件已保存，大小: {os.path.getsize(dest)} 字节")

        ext = Path(uf.name).suffix.lower()
        add_debug(f"文件扩展名: '{ext}'")

        if ext in (".xlsx", ".xls") and not uf.name.startswith("~$"):
            add_debug(f"✓ 扩展名匹配，尝试添加到解析列表")
            try:
                if ext == ".xls" and _XLRD_OK:
                    add_debug(f"尝试用xlrd打开文件...")
                    book = xlrd.open_workbook(dest)
                    add_debug(f"  ✓ 成功打开，sheet数量: {book.nsheets}")
                paths.append((Path(uf.name).stem, dest))
                add_debug(f"成功添加文件: {uf.name}")
            except Exception as e:
                add_debug(f"✗ 文件验证失败: {e}")
                st.warning(f"文件 {uf.name} 可能已损坏，将跳过")
        else:
            add_debug(f"✗ 文件被跳过: 扩展名不支持或是临时文件")

    add_debug(f"\n总共添加了 {len(paths)} 个文件到解析列表")
    add_debug("=" * 50)
    return tmp, paths


# ─────────────────────────────────────────────────────────────────────────────
# 主程序
# ─────────────────────────────────────────────────────────────────────────────
def main():
    # 修改问题1：去掉调试横线，改为更简洁的调试信息
    debug_messages = []
    debug_placeholder = st.empty()  # 用于显示调试信息的占位符

    def add_debug(msg):
        """添加调试信息 - 简化版，去掉横线"""
        # 只在调试模式下显示简短信息
        if st.session_state.get("debug_mode", False):
            st.caption(f"🔍 {msg}")  # 使用 caption 更不显眼
        # 仍然打印到终端，但去掉横线
        print(msg)

    st.markdown('<p class="main-header">📅 班级课表空余时间统计分析</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">异步解析 · 磁盘缓存 · 懒加载图表 · 多格式导出</p>', unsafe_allow_html=True)

    # 添加调试模式开关（默认关闭）
    with st.sidebar:
        st.markdown("---")
        debug_mode = st.checkbox("🔧 显示调试信息", value=False, help="开启后显示解析过程信息")
        st.session_state.debug_mode = debug_mode

    # 只在开启调试模式时显示调试区域
    if debug_mode:
        st.markdown("### 🔧 调试信息")
    # debug_placeholder 已经在上面定义了，这里不需要重复
    # ── 侧边栏 ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## 📂 数据上传")
        upload_mode = st.radio(
            "上传方式",
            ["ZIP 压缩包", "多文件（文件夹/批量）"],
            horizontal=True,
        )

        uploaded_zip = None
        uploaded_files = None

        if upload_mode == "ZIP 压缩包":
            uploaded_zip = st.file_uploader(
                "上传 ZIP 文件（内含课表 XLSX / XLS）",
                type=["zip"],
                help="可将整个文件夹压缩为 ZIP 后上传，支持中文路径",
            )
        else:
            uploaded_files = st.file_uploader(
                "批量选择课表文件（XLSX / XLS，可多选）",
                type=["xlsx", "xls"],
                accept_multiple_files=True,
                help="可框选整个文件夹中的 xlsx / xls 文件后一次性上传",
            )

        st.markdown("---")
        st.markdown("## ⚙️ 解析设置")
        max_workers = st.slider("并行线程数", 1, 8, 4, help="增加可提升速度，但占用更多内存")
        show_issues = st.checkbox("显示解析问题", value=True)

        st.markdown("---")
        st.markdown("## 📁 导出设置")

        # 添加导出位置选项（新增）
        export_location = st.radio(
            "导出文件保存位置",
            ["📥 浏览器下载（默认）", "💾 服务器临时目录（需要手动下载）"],
            help="选择文件保存位置。浏览器下载会自动保存到您的下载文件夹，服务器临时目录需要点击链接下载"
        )

        st.markdown("---")
        st.markdown("## ℹ️ 使用说明")
        with st.expander("查看说明"):
            st.markdown("""
**上传方式**
- **ZIP**：将课表文件夹压缩为 `.zip` 上传，支持子目录
- **多文件**：直接拖入多个 `.xlsx` 或 `.xls` 文件，或整个文件夹

> ⚠️ **读取 .xls 需安装 xlrd**
> ```
> pip install xlrd>=2.0.1
> ```

**支持的时间格式**


**导出格式**
- 📄 Markdown 报告
- 📊 Excel（多 Sheet）- 包含每周详细数据
- 📦 CSV 打包 ZIP
- 🖼️ 热力图 PNG
- 📑 趋势图 PNG

**导出位置说明**
- **浏览器下载**：文件自动保存到浏览器的默认下载文件夹
- **服务器临时目录**：文件保存在服务器临时目录，通过链接手动下载
            """)

    # ── 主内容 ───────────────────────────────────────────────────────────────
    has_input = (uploaded_zip is not None) or (uploaded_files and len(uploaded_files) > 0)

    if not has_input:
        st.markdown("""
        <div style="text-align:center;padding:4rem 1rem;color:#9CA3AF;">
            <div style="font-size:3rem;margin-bottom:1rem;">📂</div>
            <div style="font-size:1.2rem;font-weight:600;color:#4B5563;">从左侧上传课表文件</div>
            <div style="margin-top:.6rem;font-size:.9rem;">
                支持 ZIP 压缩包 或 直接拖入多个 XLSX / XLS 文件
            </div>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── 解压 / 收集文件 ──────────────────────────────────────────────────────
    temp_dir = None
    try:
        with st.spinner("📂 收集文件中…"):
            if uploaded_zip is not None:
                temp_dir, all_paths = collect_xlsx_from_zip(uploaded_zip, add_debug)
            else:
                temp_dir, all_paths = collect_xlsx_from_files(uploaded_files, add_debug)

        if not all_paths:
            st.error("未找到任何有效的 .xlsx / .xls 课表文件，请检查上传内容。")
            return

        c1, c2, c3 = st.columns(3)
        c1.metric("📄 发现课表文件", len(all_paths))
        c2.metric("⚡ 解析线程", max_workers)
        c3.metric("🗄️ 缓存目录", str(CACHE_DIR)[-30:])

        # ── 异步并行解析 ─────────────────────────────────────────────────────
        prog_bar  = st.progress(0)
        prog_text = st.empty()

        def _cb(ratio, msg):
            prog_bar.progress(ratio)
            prog_text.text(msg)

        t0 = time.time()
        all_occ, all_issues = run_parse(all_paths, max_workers, _cb)
        elapsed = time.time() - t0

        prog_bar.empty()
        prog_text.empty()
        st.success(f"✅ 成功解析 **{len(all_occ)}** 份课表，耗时 {elapsed:.2f}s")

        if not all_occ:
            st.error("没有成功解析的课表，请检查文件格式。")
            return

        if show_issues and all_issues:
            with st.expander(f"⚠️ 解析问题（{len(all_issues)} 份）"):
                for nm, iss_list in all_issues.items():
                    for iss in iss_list:
                        st.warning(f"**{nm}**: {iss}")

        # ── 计算 free4d（稀疏聚合） ──────────────────────────────────────────
        free4d, n = build_free4d(all_occ)

        # ── KPI 概览 ─────────────────────────────────────────────────────────
        st.markdown('<p class="section-title">📊 数据概览</p>', unsafe_allow_html=True)
        k1, k2, k3, k4 = st.columns(4)
        total_slots = int(sum(occ.sum() for occ in all_occ.values()))
        free_rate   = (1 - total_slots / (n * TOTAL_WEEKS * 7 * 12)) * 100
        fully_free  = int((free4d == n).sum())
        avg_free_pw = free4d.mean(axis=(1,2)).mean()

        for col, (val, label, delta) in zip(
            [k1, k2, k3, k4],
            [
                (n,                "成功解析人数",   ""),
                (f"{free_rate:.1f}%", "总体空闲率",   ("🟢 较高" if free_rate>60 else "🟡 中等" if free_rate>40 else "🔴 偏低")),
                (fully_free,       "全员空闲节次总数", ""),
                (f"{avg_free_pw:.1f}", "周均空闲人次",  ""),
            ],
        ):
            col.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-value">{val}</div>
                <div class="kpi-label">{label}</div>
                <div class="kpi-delta">{delta}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("")

        # ── 标签页（懒加载：仅在首次切换到该 Tab 时计算） ────────────────────
        if "rendered_tabs" not in st.session_state:
            st.session_state.rendered_tabs = set()

        tab_labels = [
            "📊 普通热力图",
            "👥 带人名热力图",
            "🧊 3D 曲面",
            "📑 导出",
        ]
        tabs = st.tabs(tab_labels)

        # Tab 0: 普通热力图
        with tabs[0]:
            st.markdown("### 普通周度热力图")
            selected_week = st.selectbox(
                "选择周次",
                range(1, TOTAL_WEEKS + 1),
                format_func=lambda x: f"第{x}周",
                key="heatmap_normal"
            )

            # 中文版
            st.plotly_chart(chart_heatmap(free4d, n, selected_week - 1, english=False),
                            use_container_width=True)

            # 英文导出
            if st.button("导出英文版", key="export_normal"):
                fig_en = chart_heatmap(free4d, n, selected_week - 1, english=True)
                img_bytes = fig_en.to_image(format="png", width=1200, height=800)

                # 根据导出位置显示不同的下载方式
                if export_location == "📥 浏览器下载（默认）":
                    st.download_button(
                        "📥 下载英文热力图",
                        data=img_bytes,
                        file_name=f"week_{selected_week}_heatmap_en.png",
                        mime="image/png",
                    )
                else:
                    # 服务器临时目录方式
                    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
                    temp_file = os.path.join(tempfile.gettempdir(), f"heatmap_{selected_week}_{timestamp}.png")
                    with open(temp_file, "wb") as f:
                        f.write(img_bytes)
                    st.success(f"✅ 文件已保存到服务器临时目录：")
                    st.code(temp_file)
                    st.info("请复制上述路径手动获取文件")

        # Tab 1: 带人名热力图
        with tabs[1]:
            st.markdown("### 带人名标注的热力图")
            st.caption("⭐ 表示全员空闲时段 | 红色格子标注缺席人员")

            threshold_names = st.slider(
                "低空闲率阈值 (%)",
                min_value=1,
                max_value=20,
                value=5,
                step=1,
                key="threshold_names",
                help="空闲率低于此值的时段会标注缺席人员"
            )

            week_names = st.selectbox(
                "选择周次",
                range(1, TOTAL_WEEKS + 1),
                format_func=lambda x: f"第{x}周",
                key="heatmap_names"
            )

            fig_names = chart_heatmap_with_names(free4d, n, all_occ, week_names - 1, threshold_names)
            st.plotly_chart(fig_names, use_container_width=True)

            # 显示全员空闲时段
            fully_free = []
            for d in range(7):
                for p in range(12):
                    if free4d[week_names - 1, d, p] == n:
                        fully_free.append({
                            "星期": DAY_NAMES[d],
                            "节次": p + 1,
                            "时间": f"{PERIOD_TIMES[p + 1][0]}–{PERIOD_TIMES[p + 1][1]}",
                        })

            if fully_free:
                st.markdown("#### ⭐ 全员空闲时段")
                st.dataframe(pd.DataFrame(fully_free), use_container_width=True)


        # Tab 2: 3D 曲面
        with tabs[2]:
            st.markdown("### 3D 空闲分布曲面")
            week_3d = st.selectbox(
                "选择周次",
                range(1, TOTAL_WEEKS + 1),
                format_func=lambda x: f"第{x}周",
                key="3d_week"
            )
            st.plotly_chart(chart_3d_surface(free4d, n, week_3d - 1), use_container_width=True)

        # Tab 3: 导出（修改问题2和问题3）
        with tabs[3]:
            st.markdown("### 批量导出")

            # 根据导出位置显示不同的提示
            if export_location == "📥 浏览器下载（默认）":
                st.info("📥 文件导出后，浏览器会自动下载到您电脑的默认下载文件夹中", icon="ℹ️")
            else:
                st.warning("💾 文件将保存到服务器临时目录，请手动复制路径获取", icon="💾")

            col1, col2 = st.columns(2)

            with col1:
                # 问题3：删除低空闲率人员分布，只保留Excel导出
                if st.button("📊 导出每周详细数据 (Excel)", use_container_width=True):
                    with st.spinner("正在生成Excel文件..."):
                        # 使用原有的build_excel_export函数，它已经包含了每周详细数据
                        excel_bytes = build_excel_export(free4d, n, all_occ)

                        st.success("✅ Excel文件生成成功！")

                        # 生成文件名
                        filename = f"课表空余分析_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                        if export_location == "📥 浏览器下载（默认）":
                            st.download_button(
                                "📥 点击下载Excel文件",
                                data=excel_bytes,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                        else:
                            # 服务器临时目录方式
                            temp_file = os.path.join(tempfile.gettempdir(), filename)
                            with open(temp_file, "wb") as f:
                                f.write(excel_bytes)
                            st.success(f"✅ 文件已保存到：")
                            st.code(temp_file)
                            st.info("请复制上述路径手动获取文件")

            with col2:
                if st.button("📦 导出所有周英文热力图", use_container_width=True):
                    import zipfile
                    zip_buffer = io.BytesIO()

                    with zipfile.ZipFile(zip_buffer, "w") as zf:
                        for w in range(TOTAL_WEEKS):
                            fig = chart_heatmap(free4d, n, w, english=True)
                            img_bytes = fig.to_image(format="png", width=1200, height=800)
                            zf.writestr(f"week_{w + 1}_heatmap_en.png", img_bytes)

                    zip_buffer.seek(0)
                    st.success("✅ 热力图打包完成！")

                    # 生成文件名
                    filename = f"heatmaps_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.zip"

                    if export_location == "📥 浏览器下载（默认）":
                        st.download_button(
                            "📥 点击下载ZIP",
                            data=zip_buffer,
                            file_name=filename,
                            mime="application/zip",
                        )
                    else:
                        # 服务器临时目录方式
                        temp_file = os.path.join(tempfile.gettempdir(), filename)
                        with open(temp_file, "wb") as f:
                            f.write(zip_buffer.getvalue())
                        st.success(f"✅ 文件已保存到：")
                        st.code(temp_file)
                        st.info("请复制上述路径手动获取文件")

            # 添加额外的导出选项
            st.markdown("---")
            st.markdown("#### 📋 其他导出格式")

            col3, col4, col5 = st.columns(3)

            with col3:
                if st.button("📄 Markdown报告", use_container_width=True):
                    report = generate_report(free4d, n, all_occ, all_issues)
                    filename = f"report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.md"

                    if export_location == "📥 浏览器下载（默认）":
                        st.download_button(
                            "📥 下载MD文件",
                            data=report,
                            file_name=filename,
                            mime="text/markdown",
                        )
                    else:
                        temp_file = os.path.join(tempfile.gettempdir(), filename)
                        with open(temp_file, "w", encoding="utf-8") as f:
                            f.write(report)
                        st.success(f"✅ 文件已保存到：")
                        st.code(temp_file)

            with col4:
                if st.button("📦 CSV压缩包", use_container_width=True):
                    csv_bytes = build_csv_zip(free4d, n, all_occ)
                    filename = f"csv_data_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.zip"

                    if export_location == "📥 浏览器下载（默认）":
                        st.download_button(
                            "📥 下载CSV-ZIP",
                            data=csv_bytes,
                            file_name=filename,
                            mime="application/zip",
                        )
                    else:
                        temp_file = os.path.join(tempfile.gettempdir(), filename)
                        with open(temp_file, "wb") as f:
                            f.write(csv_bytes)
                        st.success(f"✅ 文件已保存到：")
                        st.code(temp_file)

            with col5:
                if st.button("📈 趋势图PNG", use_container_width=True):
                    trend_bytes = mpl_trend_bytes(free4d, n)
                    filename = f"trend_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.png"

                    if export_location == "📥 浏览器下载（默认）":
                        st.download_button(
                            "📥 下载趋势图",
                            data=trend_bytes,
                            file_name=filename,
                            mime="image/png",
                        )
                    else:
                        temp_file = os.path.join(tempfile.gettempdir(), filename)
                        with open(temp_file, "wb") as f:
                            f.write(trend_bytes)
                        st.success(f"✅ 文件已保存到：")
                        st.code(temp_file)

            # 添加服务器临时目录文件管理功能
            if export_location == "💾 服务器临时目录（需要手动下载）":
                st.markdown("---")
                st.markdown("#### 📁 已保存的文件")

                # 列出临时目录中的相关文件
                temp_files = []
                for f in os.listdir(tempfile.gettempdir()):
                    if f.startswith(("课表空余分析", "heatmaps_", "report_", "csv_data_", "trend_")):
                        file_path = os.path.join(tempfile.gettempdir(), f)
                        temp_files.append({
                            "文件名": f,
                            "大小": f"{os.path.getsize(file_path) / 1024:.1f} KB",
                            "修改时间": time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(file_path)))
                        })

                if temp_files:
                    st.dataframe(pd.DataFrame(temp_files), use_container_width=True)
                    if st.button("🗑️ 清空临时文件", use_container_width=True):
                        for f in os.listdir(tempfile.gettempdir()):
                            if f.startswith(("课表空余分析", "heatmaps_", "report_", "csv_data_", "trend_")):
                                try:
                                    os.remove(os.path.join(tempfile.gettempdir(), f))
                                except:
                                    pass
                        st.success("临时文件已清空")
                        st.rerun()
                else:
                    st.info("暂无已保存的文件")

    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    main()